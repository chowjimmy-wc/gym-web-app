import openpyxl
from pathlib import Path

path = Path(__file__).parent / 'Complete_60Days_Lean_Bulk_Plan.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print('=' * 70)
    print(f'SHEET: {name}  ({ws.max_row} rows x {ws.max_column} cols)')
    print('=' * 70)
    for row in ws.iter_rows(values_only=True):
        cells = ['' if c is None else str(c) for c in row]
        if any(cells):
            print(' | '.join(cells))
    print()
