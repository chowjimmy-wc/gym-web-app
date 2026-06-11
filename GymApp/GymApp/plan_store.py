"""Plan data storage, Excel import, and CRUD helpers."""
from __future__ import annotations

import json
import re
from io import BytesIO
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
DATA = ROOT / "data"
PLAN_FILE = DATA / "plan.json"

SHEET_NUTRITION = "1.營養與熱量目標"
SHEET_WORKOUT = "2.60天詳細課表"
SHEET_MEALS = "3.60天詳細餐單"
SHEET_PROGRESS = "4.進度追蹤與檢討"

DEFAULT_SHEET_ORDER = [SHEET_NUTRITION, SHEET_WORKOUT, SHEET_MEALS, SHEET_PROGRESS]

SECTION_MAP = {
    "nutrition": SHEET_NUTRITION,
    "workouts": SHEET_WORKOUT,
    "meals": SHEET_MEALS,
    "progress": SHEET_PROGRESS,
}


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def day_num(day_str: str) -> int:
    m = re.search(r"(\d+)", day_str or "")
    return int(m.group(1)) if m else 0


def read_table(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_str(h) for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(row):
            continue
        record = {headers[i]: cell_str(row[i]) for i in range(len(headers))}
        if any(record.values()):
            records.append(record)
    return records


def import_workbook(source: Path | bytes, title: str | None = None) -> dict:
    if isinstance(source, bytes):
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)

    plan = {
        "title": title or "Imported Plan",
        "sheets": {name: read_table(wb[name]) for name in wb.sheetnames},
        "sheet_order": list(wb.sheetnames),
    }
    return plan


def default_plan() -> dict:
    return {
        "title": "New Gym Plan",
        "sheets": {name: [] for name in DEFAULT_SHEET_ORDER},
        "sheet_order": list(DEFAULT_SHEET_ORDER),
    }


def load_plan() -> dict:
    if PLAN_FILE.exists():
        return json.loads(PLAN_FILE.read_text(encoding="utf-8"))
    return default_plan()


def save_plan(plan: dict) -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    PLAN_FILE.write_text(
        json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def sheet_name(section: str) -> str:
    if section not in SECTION_MAP:
        raise KeyError(f"Unknown section: {section}")
    return SECTION_MAP[section]


def get_records(plan: dict, section: str) -> list[dict]:
    name = sheet_name(section)
    if name not in plan["sheets"]:
        plan["sheets"][name] = []
    return plan["sheets"][name]


def max_day(plan: dict) -> int:
    days = [day_num(r.get("天數", "")) for r in get_records(plan, "workouts")]
    return max(days) if days else 0


def find_by_day(records: list[dict], day: int) -> int | None:
    for i, r in enumerate(records):
        if day_num(r.get("天數", "")) == day:
            return i
    return None


def query_records(records: list[dict], q: str = "", day: int | None = None) -> list[dict]:
    results = []
    for i, r in enumerate(records):
        if day is not None and day_num(r.get("天數", "")) != day:
            continue
        if q:
            text = " ".join(str(v) for v in r.values()).lower()
            if q.lower() not in text:
                continue
        results.append({"index": i, **r})
    return results


def create_record(plan: dict, section: str, record: dict) -> dict:
    records = get_records(plan, section)
    if section in ("workouts", "meals"):
        d = day_num(record.get("天數", ""))
        if d and find_by_day(records, d) is not None:
            raise ValueError(f"Day {d} already exists")
    records.append(record)
    save_plan(plan)
    return record


def update_record(plan: dict, section: str, index: int, record: dict) -> dict:
    records = get_records(plan, section)
    if index < 0 or index >= len(records):
        raise IndexError("Record not found")
    records[index] = record
    save_plan(plan)
    return record


def upsert_by_day(plan: dict, section: str, day: int, record: dict) -> dict:
    records = get_records(plan, section)
    record["天數"] = f"Day {day}"
    idx = find_by_day(records, day)
    if idx is None:
        records.append(record)
    else:
        records[idx] = {**records[idx], **record}
    save_plan(plan)
    return record


def delete_record(plan: dict, section: str, index: int) -> None:
    records = get_records(plan, section)
    if index < 0 or index >= len(records):
        raise IndexError("Record not found")
    records.pop(index)
    save_plan(plan)


def delete_by_day(plan: dict, section: str, day: int) -> None:
    records = get_records(plan, section)
    idx = find_by_day(records, day)
    if idx is None:
        raise IndexError("Record not found")
    records.pop(idx)
    save_plan(plan)
